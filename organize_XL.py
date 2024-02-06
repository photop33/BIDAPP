import openpyxl
from datetime import datetime

# Load the Excel workbook
workbook = openpyxl.load_workbook("extracted_data.xlsx")

# Select the active sheet
sheet = workbook.active

# Create a list to store the row numbers to delete
rows_to_delete = []

# Iterate over all rows in column A starting from row 2
for row_number, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
    print(row, row_number)
    # Check if the cell value is not None
    if row[0] is not None:
        try:
            # Attempt to parse the cell value as a date
            date_value = datetime.strptime(row[0], "%d/%m/%Y")
            # If successful, replace the cell value with the datetime object
            sheet.cell(row=row_number, column=1, value=date_value)
        except ValueError:
            # If parsing fails, mark the row for deletion
            rows_to_delete.append(row_number)
    else:
        # If the cell value is None, mark the row for deletion
        rows_to_delete.append(row_number)

# Delete rows marked for deletion (starting from the last row to avoid shifting issues)
for row_number in reversed(rows_to_delete):
    sheet.delete_rows(row_number)

# Add header row
sheet.insert_rows(1)
# Set column headers
sheet.cell(row=1, column=1, value="Date")
sheet.cell(row=1, column=2, value="Description")
sheet.cell(row=1, column=3, value="Amount")
sheet.delete_rows(2)
# Save the modified workbook
workbook.save("modified_excel_file.xlsx")
